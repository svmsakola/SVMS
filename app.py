import os
import io
import sys
import uuid
import json
import random
import signal
import base64
import traceback
import time
import threading
from datetime import datetime, timedelta
from functools import wraps
import bcrypt
import cv2
import numpy as np
import face_recognition
from PIL import Image
from flask import (
    Flask,
    render_template,
    request,
    redirect,
    url_for,
    session,
    jsonify,
    send_file,
    flash,
    make_response,
)
from flask_cors import CORS
import openpyxl
from openpyxl.styles import Font, Alignment
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from ultralytics import YOLO
from werkzeug.utils import secure_filename
import boto3
from botocore.exceptions import ClientError

# Try torch globally so we can use torch.no_grad() in detect_face
try:
    import torch
except Exception:
    torch = None

app = Flask(__name__)
CORS(app)
app.secret_key = os.urandom(24)

# Base prefixes (kept short) — keys will be constructed via s3_key(...)
UPLOAD_PREFIX = "uploads/docimg"
PDF_PREFIX = "pdf"
REMARK_PDF_PREFIX = "remarkpdf"
ENCODINGS_PREFIX = "encodings"
FACEDATA_PREFIX = "facedata"
PROFILES_PREFIX = "profiles"
# Optional top-level prefix if you want all keys under a single folder in the bucket:
S3_BASE_PREFIX = os.environ.get("S3_BASE_PREFIX", "").strip("/")  # e.g. "svms" or "" for none

FACE_TABLE = os.environ.get("FACE_TABLE", "SVMSFaceData")
DEPT_TABLE = os.environ.get("DEPT_TABLE", "SVMSDepartments")
AUTH_TABLE = os.environ.get("AUTH_TABLE", "SVMSAuth")
BUCKET = os.environ.get("S3_BUCKET", "svmsakola")
AWS_REGION = os.environ.get("AWS_REGION", "ap-south-1")
ALLOWED_EXTENSIONS = {"pdf"}

dynamodb = boto3.resource("dynamodb", region_name=AWS_REGION)
s3 = boto3.client("s3", region_name=AWS_REGION)


def s3_key(*parts):
    """
    Build a clean S3 key by joining provided parts with '/' and stripping extra slashes.
    If S3_BASE_PREFIX is set, it will be prefixed automatically.
    """
    parts_clean = [str(p).strip("/") for p in parts if p is not None and str(p).strip() != ""]
    if S3_BASE_PREFIX:
        return "/".join([S3_BASE_PREFIX] + parts_clean)
    return "/".join(parts_clean)


def s3_put_bytes(key, b, content_type=None):
    kwargs = {"Bucket": BUCKET, "Key": key, "Body": b}
    if content_type:
        kwargs["ContentType"] = content_type
    s3.put_object(**kwargs)
    return True


def s3_get_bytes(key):
    try:
        res = s3.get_object(Bucket=BUCKET, Key=key)
        return res["Body"].read()
    except ClientError as e:
        if e.response["Error"]["Code"] == "NoSuchKey":
            raise FileNotFoundError
        raise


def s3_exists(key):
    try:
        s3.head_object(Bucket=BUCKET, Key=key)
        return True
    except ClientError:
        return False


def s3_delete(key):
    try:
        s3.delete_object(Bucket=BUCKET, Key=key)
        return True
    except Exception:
        return False


def allowed_file(filename):
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_EXTENSIONS


def signal_handler(sig, frame):
    sys.exit(0)


signal.signal(signal.SIGINT, signal_handler)
signal.signal(signal.SIGTERM, signal_handler)

os.makedirs("/tmp", exist_ok=True)

if sys.platform.startswith("darwin"):
    cv2.setNumThreads(4)
    os.environ["OPENCV_OPENCL_RUNTIME"] = ""
elif sys.platform.startswith("linux"):
    cv2.setNumThreads(0)
    try:
        if cv2.ocl.haveOpenCL():
            cv2.ocl.setUseOpenCL(True)
    except Exception:
        pass

# Decide which face_recognition model to use
face_recognition_model = "hog"
if sys.platform.startswith("linux"):
    try:
        if torch and torch.cuda.is_available():
            face_recognition_model = "cnn"
    except Exception:
        pass

# Load YOLO models
try:
    person_model = YOLO("models/yolo11n-seg.pt")
    face_model = YOLO("models/yolov11n-face.pt")
except Exception as e:
    print(f"Error loading YOLO models: {e}")
    sys.exit(1)

# ------------------------------------------------------------------
# FAST FACE RECOGNITION CACHE (GLOBAL SHARED STATE PER WORKER)
# ------------------------------------------------------------------

FACE_CACHE_TTL = 60  # seconds to keep cache "fresh"

_face_cache_lock = threading.RLock()
_face_cache = {
    "encodings": None,   # np.ndarray of shape (N,128) float32
    "uids":     [],      # [uid, uid, ...] parallel to encodings rows
    "last_load": 0.0     # unix timestamp
}


def _load_face_cache(force: bool = False):
    """
    Refresh the global in-memory known-face cache from DynamoDB+S3.

    We only rebuild if:
      - cache empty,
      - cache is older than FACE_CACHE_TTL,
      - or force=True.

    This dramatically reduces repeated S3/DB lookups per request.
    NOTE: We do NOT regenerate encodings from raw images here
    (that "healing" logic in your old /detect_face was expensive and
    caused per-request slowness). That should be handled offline.
    """
    global _face_cache

    with _face_cache_lock:
        now = time.time()

        # Use cached version if still fresh
        if (
            not force
            and _face_cache["encodings"] is not None
            and (now - _face_cache["last_load"] < FACE_CACHE_TTL)
        ):
            return

        # Pull all visitors from DB
        known_data = scan_face_data()

        enc_list = []
        uid_list = []

        for uid, user in known_data.items():
            # Use precomputed encodings if available
            for encoding_path in user.get("face_encodings", []):
                try:
                    raw = s3_get_bytes(encoding_path)  # load .npy bytes
                    arr = np.load(io.BytesIO(raw))
                    # Expect shape (128,), skip bad arrays
                    if arr is not None and hasattr(arr, "shape") and arr.shape == (128,):
                        enc_list.append(arr.astype("float32"))
                        uid_list.append(uid)
                except Exception as e:
                    print(f"[face_cache] Error loading {encoding_path} for {uid}: {e}")
                    continue

        if enc_list:
            enc_array = np.stack(enc_list, axis=0).astype("float32")  # (N,128)
        else:
            enc_array = np.zeros((0, 128), dtype="float32")

        _face_cache["encodings"] = enc_array
        _face_cache["uids"] = uid_list
        _face_cache["last_load"] = now


def _match_face(embedding: np.ndarray, tolerance: float = 0.4):
    """
    Fast approximate nearest-neighbor match.

    We compute L2 distance between the captured embedding (128-d)
    and each cached encoding, then pick the closest.

    If best distance <= tolerance, we return the UID, else None.
    """
    with _face_cache_lock:
        encs = _face_cache["encodings"]
        uids = _face_cache["uids"]

        if encs is None or encs.shape[0] == 0:
            return None

        # Vectorized distance: (N,128) vs (1,128)
        diff = encs - embedding.astype("float32")
        dists = np.linalg.norm(diff, axis=1)  # shape (N,)

        best_idx = int(np.argmin(dists))
        best_dist = float(dists[best_idx])

        # Your old tolerance via compare_faces was 0.4
        if best_dist <= tolerance:
            return uids[best_idx]
        else:
            return None


def cleanup_memory():
    import gc

    gc.collect()
    if sys.platform.startswith("linux"):
        try:
            import ctypes

            libc = ctypes.CDLL("libc.so.6")
            libc.malloc_trim(0)
        except Exception:
            pass


def generate_visitor_id():
    today = datetime.now()
    date_component = today.strftime("%Y%m%d")
    table = dynamodb.Table(FACE_TABLE)
    try:
        resp = table.scan(ProjectionExpression="uid,visitor")
    except Exception:
        return f"V{date_component}0001"
    items = resp.get("Items", [])
    today_visitors = []
    for it in items:
        for visit in it.get("visitor", []):
            vid = visit.get("visit_id")
            if vid and vid.startswith(f"V{date_component}"):
                today_visitors.append(vid)
    visit_count = len(today_visitors) + 1
    return f"V{date_component}{str(visit_count).zfill(4)}"


def is_marathi(text):
    if not text:
        return False
    return any("\u0900" <= c <= "\u097F" for c in text)


def get_face_item(uid):
    table = dynamodb.Table(FACE_TABLE)
    try:
        resp = table.get_item(Key={"uid": uid})
        return resp.get("Item")
    except Exception:
        return None


def put_face_item(uid, data):
    table = dynamodb.Table(FACE_TABLE)
    data_copy = dict(data)
    data_copy["uid"] = uid
    table.put_item(Item=data_copy)
    return True


def scan_face_data():
    table = dynamodb.Table(FACE_TABLE)
    data = {}
    try:
        resp = table.scan()
        items = resp.get("Items", [])
        while "LastEvaluatedKey" in resp:
            resp = table.scan(ExclusiveStartKey=resp["LastEvaluatedKey"])
            items.extend(resp.get("Items", []))
        for it in items:
            uid = it.get("uid")
            if uid:
                data[uid] = it
    except Exception:
        pass
    return data


def delete_face_item(uid):
    table = dynamodb.Table(FACE_TABLE)
    try:
        table.delete_item(Key={"uid": uid})
        return True
    except Exception:
        return False


def load_users():
    table = dynamodb.Table(AUTH_TABLE)
    users = {}
    try:
        resp = table.scan()
        items = resp.get("Items", [])
        while "LastEvaluatedKey" in resp:
            resp = table.scan(ExclusiveStartKey=resp["LastEvaluatedKey"])
            items.extend(resp.get("Items", []))
        for it in items:
            users[it["username"]] = it
        return users
    except Exception:
        return {}


def save_users(users):
    table = dynamodb.Table(AUTH_TABLE)
    for username, u in users.items():
        item = {"username": username, "password": u["password"], "role": u["role"]}
        table.put_item(Item=item)


def save_face_data(data):
    table = dynamodb.Table(FACE_TABLE)
    try:
        for uid, ud in data.items():
            item = dict(ud)
            item["uid"] = uid
            table.put_item(Item=item)
        return True
    except Exception:
        return False


def save_face_images_to_s3(frame, uid):
    results = face_model.predict(source=frame, stream=True)
    saved_images = []
    face_encodings = []
    for r in results:
        if r.boxes is not None:
            for box in r.boxes.data:
                x1, y1, x2, y2 = map(int, box[:4].cpu().numpy())
                face_crop = frame[y1:y2, x1:x2]
                if face_crop.size == 0:
                    continue
                _, buf = cv2.imencode(".jpg", face_crop)
                b = buf.tobytes()
                filename = s3_key(FACEDATA_PREFIX, uid, f"{uid}_img{len(saved_images) + 1}.jpg")
                s3_put_bytes(filename, b, "image/jpeg")
                saved_images.append(filename)
                rgb_face = cv2.cvtColor(face_crop, cv2.COLOR_BGR2RGB)
                encs = face_recognition.face_encodings(rgb_face, model=face_recognition_model)
                if encs:
                    arr = encs[0]
                    bio = io.BytesIO()
                    np.save(bio, arr)
                    bio.seek(0)
                    encoding_filename = s3_key(ENCODINGS_PREFIX, f"{uid}_encoding{len(face_encodings) + 1}.npy")
                    s3_put_bytes(encoding_filename, bio.read(), "application/octet-stream")
                    face_encodings.append(encoding_filename)
                if len(saved_images) >= 3:
                    break
    return saved_images, face_encodings


def detect_person_and_face(frame):
    white_bg = np.ones_like(frame) * 255
    frame_area = frame.shape[0] * frame.shape[1]
    results = person_model.predict(source=frame, stream=True, classes=[0])
    max_area = 0
    selected_mask = None
    for r in results:
        if r.boxes is not None and r.masks is not None:
            for box, mask in zip(r.boxes.data, r.masks.data):
                conf = float(box[4].cpu().numpy())
                x1, y1, x2, y2 = map(int, box[:4].cpu().numpy())
                box_area = (x2 - x1) * (y2 - y1)
                if box_area / frame_area >= 0.15 and conf >= 0.5 and box_area > max_area:
                    max_area = box_area
                    selected_mask = mask.cpu().numpy().astype("float32")
    face_output = white_bg.copy()
    face_crop = None
    if selected_mask is not None:
        selected_mask = cv2.resize(selected_mask, (frame.shape[1], frame.shape[0]))
        _, binary_mask = cv2.threshold(selected_mask, 0.5, 255, cv2.THRESH_BINARY)
        binary_mask = cv2.merge([binary_mask.astype(np.uint8)] * 3)
        person_only = cv2.bitwise_and(frame, binary_mask)
        person_with_white_bg = np.where(binary_mask == 0, white_bg, person_only)
        face_results = face_model.predict(source=person_with_white_bg, stream=True)
        max_face_area = 0
        selected_face = None
        for fr in face_results:
            if fr.boxes is not None:
                for box in fr.boxes.data:
                    x1, y1, x2, y2 = map(int, box[:4].cpu().numpy())
                    face_area = (x2 - x1) * (y2 - y1)
                    if face_area > max_face_area:
                        max_face_area = face_area
                        selected_face = (x1, y1, x2, y2)
        if selected_face:
            x1, y1, x2, y2 = selected_face
            w, h = x2 - x1, y2 - y1
            pad_w, pad_h = int(w * 0.6), int(h * 0.4)
            x1, y1 = max(x1 - pad_w, 0), max(y1 - pad_h, 0)
            x2, y2 = min(x2 + pad_w, frame.shape[1]), min(y2 + pad_h, frame.shape[0])
            face_crop = person_with_white_bg[y1:y2, x1:x2]
            face_output[y1:y2, x1:x2] = face_crop
    return face_output, face_crop


@app.route("/", methods=["GET", "POST"])
def login():
    if "user" in session:
        return redirect(url_for("dashboard", role=session["user"]["role"]))
    if request.method == "POST":
        if request.is_json:
            data = request.json
            if not data:
                return jsonify(success=False, message="Invalid request format")
            username = data.get("username")
            password = data.get("password")
            if not username or not password:
                return jsonify(success=False, message="Username and password are required")
            users = load_users()
            if (
                username in users
                and bcrypt.checkpw(password.encode("utf-8"), users[username]["password"].encode("utf-8"))
            ):
                role = users[username]["role"]
                session["user"] = {"username": username, "role": role}
                return jsonify(success=True, redirect_url=url_for("dashboard", role=role))
            return jsonify(success=False, message="Invalid credentials")
        else:
            username = request.form.get("username")
            password = request.form.get("password")
            if not username or not password:
                flash("Username and password are required", "error")
                return render_template("login.html")
            users = load_users()
            if (
                username in users
                and bcrypt.checkpw(password.encode("utf-8"), users[username]["password"].encode("utf-8"))
            ):
                role = users[username]["role"]
                session["user"] = {"username": username, "role": role}
                return redirect(url_for("dashboard", role=role))
            flash("Invalid credentials", "error")
            return render_template("login.html")
    requested_path = request.path
    if requested_path == "/":
        return render_template("index.html")
    else:
        return render_template("login.html")


@app.route("/login", methods=["GET", "POST"])
def login_page():
    return login()


@app.route("/sw/sw.js")
def service_worker():
    return app.send_static_file("sw/sw.js")


@app.route("/register", methods=["GET", "POST"])
def register():
    if request.method == "POST":
        data = request.json
        if not data:
            return jsonify(success=False, message="Invalid request format")
        username = data.get("username")
        password = data.get("password")
        role = data.get("role")
        pin = data.get("pin")
        if not username or not password or not role or not pin:
            return jsonify(success=False, message="All fields are required")
        if pin != "100444":
            return jsonify(success=False, message="Invalid security PIN")
        valid_roles = ["admin", "registrar", "inoffice", "cctv", "scanner"]
        if role not in valid_roles:
            return jsonify(success=False, message="Invalid role selected")
        users = load_users()
        if username in users:
            return jsonify(success=False, message="Username already exists")
        hashed_password = bcrypt.hashpw(password.encode(), bcrypt.gensalt()).decode()
        users[username] = {"password": hashed_password, "role": role}
        save_users(users)
        table = dynamodb.Table(AUTH_TABLE)
        table.put_item(Item={"username": username, "password": hashed_password, "role": role})
        return jsonify(success=True)
    return render_template("register.html")


@app.route("/dashboard/<role>")
def dashboard(role):
    if "user" not in session or session["user"]["role"] != role:
        return redirect(url_for("login"))
    templates = {
        "admin": "admin/admin.html",
        "registrar": "dashboard/registrar.html",
        "inoffice": "dashboard/inoffice.html",
        "cctv": "dashboard/cctv.html",
        "scanner": "dashboard/scanner.html",
    }
    if role not in templates:
        return render_template("index.html")
    return render_template(templates[role])


@app.route("/logout")
def logout():
    session.pop("user", None)
    return redirect(url_for("login"))


@app.route("/register_visitor", methods=["POST"])
def register_visitor():
    data = request.get_json()
    if not data:
        return jsonify({"error": "Invalid request, no data received"}), 400
    if not data.get("name") or not data.get("phone"):
        missing = []
        if not data.get("name"):
            missing.append("name")
        if not data.get("phone"):
            missing.append("phone")
        return jsonify({"error": f"Missing required fields: {', '.join(missing)}"}), 400

    try:
        facedata = scan_face_data()
    except Exception:
        facedata = {}

    existing_uid = None
    for uid_key, user_data_val in facedata.items():
        if isinstance(user_data_val, dict):
            if (
                user_data_val.get("name") == data.get("name")
                and user_data_val.get("phone") == data.get("phone")
            ):
                existing_uid = uid_key
                break
        else:
            print(f"Warning: Skipping non-dict entry for UID {uid_key} in facedata")

    uid = existing_uid if existing_uid else f"UID{int(datetime.now().timestamp())}"

    tahasil_options = ["अकोला", "अकोट", "तेल्हरा", "बाळापूर", "पातूर", "मुर्तिजापूर", "बार्शीटाकळी"]
    selected_tahasil = data.get("tahasil", "")
    if selected_tahasil.endswith("<"):
        selected_tahasil = selected_tahasil[:-1]
    if selected_tahasil and selected_tahasil not in tahasil_options:
        return jsonify({"error": f"Invalid tahasil selection: {selected_tahasil}"}), 400

    visitor_data = {
        "name": data.get("name", ""),
        "phone": data.get("phone", ""),
        "email": data.get("email", ""),
        "address": data.get("address", ""),
        "tahasil": selected_tahasil,
        "district": data.get("district", "Akola"),
    }

    if not existing_uid:
        visitor_data["registration_datetime"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        if "frame" in data and data["frame"]:
            try:
                frame_data = data["frame"]
                if "," in frame_data:
                    frame_data = frame_data.split(",")[1]
                frame_bytes = base64.b64decode(frame_data)
                nparr = np.frombuffer(frame_bytes, np.uint8)
                frame = cv2.imdecode(nparr, cv2.IMREAD_COLOR)
                if frame is not None:
                    saved_images, face_encodings = save_face_images_to_s3(frame, uid)
                    if saved_images:
                        visitor_data["images"] = saved_images
                    if face_encodings:
                        visitor_data["face_encodings"] = face_encodings
                else:
                    return jsonify({"error": "Invalid frame data, unable to decode"}), 400
            except Exception as e:
                return jsonify({"error": f"Error processing frame: {str(e)}"}), 500
            finally:
                cleanup_memory()

    visit_status = "पुनरावृत्ती अभ्यागत" if existing_uid else "नवीन अभ्यागत"
    visitor_id = generate_visitor_id()
    visit_entry = {
        "visit_id": visitor_id,
        "datetime": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "purpose": data.get("purpose", ""),
        "status": visit_status,
    }

    if existing_uid and get_face_item(existing_uid):
        user_record = get_face_item(existing_uid)
        user_record["name"] = visitor_data["name"]
        user_record["phone"] = visitor_data["phone"]
        if visitor_data.get("email"):
            user_record["email"] = visitor_data["email"]
        if visitor_data.get("address"):
            user_record["address"] = visitor_data["address"]
        if visitor_data.get("tahasil"):
            user_record["tahasil"] = visitor_data["tahasil"]
        if visitor_data.get("district"):
            user_record["district"] = visitor_data["district"]
        if "visitor" not in user_record or not isinstance(user_record.get("visitor"), list):
            user_record["visitor"] = []
        user_record["visitor"].append(visit_entry)
        if "images" in visitor_data:
            user_record.setdefault("images", []).extend(visitor_data["images"])
        if "face_encodings" in visitor_data:
            user_record.setdefault("face_encodings", []).extend(visitor_data["face_encodings"])
        put_face_item(existing_uid, user_record)
    else:
        visitor_data["visitor"] = [visit_entry]
        put_face_item(uid, visitor_data)

    return jsonify({"success": True, "message": "Visitor registered successfully", "uid": uid, "visit_id": visitor_id})


@app.route("/search_visitors", methods=["GET"])
def search_visitors():
    query = request.args.get("query", "").lower()
    if not query:
        return jsonify({"success": False, "message": "Search query is required"})
    facedata = scan_face_data()
    matching_visitors = []
    for uid, visitor_data in facedata.items():
        if (
            query in visitor_data.get("name", "").lower()
            or query in uid.lower()
            or query in visitor_data.get("phone", "").lower()
            or any(query in visit.get("visit_id", "").lower() for visit in visitor_data.get("visitor", []))
        ):
            last_visit = visitor_data.get("visitor", [])[-1] if visitor_data.get("visitor") else None
            matching_visitors.append(
                {
                    "uid": uid,
                    "name": visitor_data.get("name", "N/A"),
                    "phone": visitor_data.get("phone", "N/A"),
                    "last_visit": last_visit.get("datetime", "N/A") if last_visit else "N/A",
                }
            )
    return jsonify({"success": True, "visitors": matching_visitors})


@app.route("/get_visitor_details", methods=["GET"])
def get_visitor_details():
    uid = request.args.get("uid")
    visit_id = request.args.get("visit_id")
    if not uid and not visit_id:
        return jsonify({"success": False, "message": "Either uid or visit_id is required"})
    if uid:
        item = get_face_item(uid)
        if item:
            return jsonify({"success": True, "visitor_data": item})
    if visit_id:
        data = scan_face_data()
        for uid, user_data in data.items():
            for visit in user_data.get("visitor", []):
                if visit.get("visit_id") == visit_id:
                    return jsonify({"success": True, "visitor_data": user_data, "uid": uid, "visit": visit})
    return jsonify({"success": False, "message": "Visitor not found"})


# ------------------------------------------------------------------
# FAST /detect_face ENDPOINT
# ------------------------------------------------------------------
@app.route("/detect_face", methods=["POST"])
def detect_face():
    # Require an uploaded frame (multipart)
    if "frame" not in request.files:
        return jsonify({"recognized": False, "message": "No frame uploaded"})

    try:
        # Read image from request
        frame_bytes = request.files["frame"].read()
        frame_arr = np.frombuffer(frame_bytes, dtype=np.uint8)
        frame = cv2.imdecode(frame_arr, cv2.IMREAD_COLOR)

        if frame is None:
            return jsonify({"recognized": False, "message": "Invalid image format"})

        # Make sure models exist
        if person_model is None or not hasattr(person_model, "predict"):
            return jsonify({"recognized": False, "message": "Model not properly initialized"})
        if face_model is None or not hasattr(face_model, "predict"):
            return jsonify({"recognized": False, "message": "Face model not initialized"})

        # Pre-resize to speed up YOLO face detection
        orig_h, orig_w = frame.shape[:2]
        target_size = 640  # YOLO-ish working size
        scale_x = orig_w / target_size
        scale_y = orig_h / target_size
        resized = cv2.resize(frame, (target_size, target_size), interpolation=cv2.INTER_LINEAR)

        # Run face detector (single forward, no grad)
        if torch:
            with torch.no_grad():
                face_results = face_model.predict(source=resized, stream=False)
        else:
            # Fallback if torch isn't available
            face_results = face_model.predict(source=resized, stream=False)

        # Pick best (highest conf) face box
        face_bbox_resized = None
        for fr in face_results:
            if fr.boxes is None or len(fr.boxes.data) == 0:
                continue

            boxes_np = fr.boxes.data.cpu().numpy()
            # YOLO boxes: [x1, y1, x2, y2, conf, ...]
            best_i = int(np.argmax(boxes_np[:, 4]))
            x1_r, y1_r, x2_r, y2_r, conf, *_ = boxes_np[best_i]

            if conf < 0.5:
                continue

            face_bbox_resized = (x1_r, y1_r, x2_r, y2_r)
            break

        if face_bbox_resized is None:
            return jsonify({"recognized": False, "message": "No face detected"})

        # Map face box back to original frame coords
        x1 = int(face_bbox_resized[0] * scale_x)
        y1 = int(face_bbox_resized[1] * scale_y)
        x2 = int(face_bbox_resized[2] * scale_x)
        y2 = int(face_bbox_resized[3] * scale_y)

        # Clamp box just in case
        x1 = max(0, min(x1, orig_w - 1))
        x2 = max(0, min(x2, orig_w - 1))
        y1 = max(0, min(y1, orig_h - 1))
        y2 = max(0, min(y2, orig_h - 1))

        if x2 <= x1 or y2 <= y1:
            return jsonify({"recognized": False, "message": "Invalid face crop"})

        # Crop the face region
        face_crop = frame[y1:y2, x1:x2]
        if face_crop.size == 0:
            return jsonify({"recognized": False, "message": "Invalid face crop"})

        # Convert to RGB for face_recognition
        rgb_face = cv2.cvtColor(face_crop, cv2.COLOR_BGR2RGB)

        # Get face embedding
        face_enc_list = face_recognition.face_encodings(
            rgb_face,
            model=face_recognition_model
        )

        if not face_enc_list:
            return jsonify({"recognized": False, "message": "Unable to encode face"})

        face_embedding = face_enc_list[0]  # (128,)

        # Refresh / load cached encodings (does nothing if cache is still warm)
        _load_face_cache(force=False)

        # Vectorized nearest neighbor vs all known encodings
        matched_uid = _match_face(face_embedding, tolerance=0.4)

        if matched_uid is None:
            # Face seen, but nobody matches under tolerance
            return jsonify({"recognized": False, "message": "Face not recognized"})

        # Get the full visitor record
        matched_user = get_face_item(matched_uid)

        return jsonify({
            "recognized": True,
            "user_data": matched_user,
            "uid": matched_uid
        })

    except Exception as e:
        return jsonify({"recognized": False, "message": f"Error processing image: {str(e)}"})

    finally:
        cleanup_memory()


@app.route("/confirm_visitor_entry", methods=["POST"])
def confirm_visitor_entry():
    data = request.json
    if not data:
        return jsonify({"success": False, "message": "Invalid request, no data received"}), 400
    visit_id = data.get("visitId")
    dvn = data.get("dvn")
    if not visit_id or not dvn:
        return jsonify({"success": False, "message": "Visit ID and DVN are required"}), 400
    try:
        data_dict = scan_face_data()
        found = False
        for uid, user_data in data_dict.items():
            for visit in user_data.get("visitor", []):
                if visit.get("visit_id") == visit_id:
                    visit["dvn"] = dvn
                    visit["entry_confirmed"] = True
                    visit["confirmation_time"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    put_face_item(uid, user_data)
                    found = True
                    break
            if found:
                break
        if not found:
            return jsonify({"success": False, "message": "Visitor not found"}), 404
        return jsonify({"success": True, "message": "Visitor entry confirmed", "dvn": dvn})
    except Exception as e:
        traceback.print_exc()
        return jsonify({"success": False, "message": f"An internal server error occurred: {str(e)}"}), 500


@app.route("/api/remove_dvn/<visit_id>", methods=["POST"])
def remove_dvn(visit_id):
    if "user" not in session:
        return jsonify({"success": False, "message": "Authentication required."}), 401
    try:
        data_dict = scan_face_data()
        found = False
        for uid, user_data in data_dict.items():
            if isinstance(user_data, dict) and "visitor" in user_data and isinstance(user_data["visitor"], list):
                for visit in user_data["visitor"]:
                    if isinstance(visit, dict) and visit.get("visit_id") == visit_id:
                        visit.pop("dvn", None)
                        visit.pop("entry_confirmed", None)
                        visit.pop("confirmation_time", None)
                        put_face_item(uid, user_data)
                        found = True
                        break
            if found:
                break
        if not found:
            return jsonify({"success": False, "message": "Visit ID not found"}), 404
        return jsonify({"success": True, "message": "DVN and entry confirmation have been removed."})
    except Exception as e:
        traceback.print_exc()
        return jsonify({"success": False, "message": f"An internal server error occurred: {str(e)}"}), 500


@app.route("/get_today_visitors", methods=["GET"])
def get_today_visitors():
    today = datetime.now().strftime("%Y-%m-%d")
    data = scan_face_data()
    today_visitors_dict = {}
    for uid, user_data in data.items():
        if isinstance(user_data, dict) and "visitor" in user_data and isinstance(user_data["visitor"], list):
            today_visits = [
                visit
                for visit in user_data.get("visitor", [])
                if isinstance(visit, dict) and visit.get("datetime", "").startswith(today)
            ]
            if today_visits:
                for visit in today_visits:
                    visit["dvn"] = visit.get("dvn")
                    visit["entry_confirmed"] = visit.get("entry_confirmed", False)
                today_visitors_dict[uid] = {
                    "name": user_data.get("name", "Unknown"),
                    "phone": user_data.get("phone", "N/A"),
                    "visitor": today_visits,
                }
    return jsonify({"success": True, "visitors": today_visitors_dict})


@app.route("/process_profile_image", methods=["POST"])
def process_profile_image():
    if "image" not in request.files:
        return jsonify({"success": False, "message": "No image uploaded"})
    file = request.files["image"]
    if file.filename == "":
        return jsonify({"success": False, "message": "No image selected"})
    uid = request.form.get("uid")
    if not uid:
        return jsonify({"success": False, "message": "User ID (uid) is required"})
    try:
        image_bytes = file.read()
        nparr = np.frombuffer(image_bytes, np.uint8)
        frame = cv2.imdecode(nparr, cv2.IMREAD_COLOR)
        if frame is None:
            return jsonify({"success": False, "message": "Invalid image format"})
        _, face_crop = detect_person_and_face(frame)
        if face_crop is None:
            return jsonify({"success": False, "message": "No face detected in the image"})
        existing_record = get_face_item(uid) or {}
        existing_images = existing_record.get("profile_img", [])
        next_img_num = len(existing_images) + 1
        _, buf = cv2.imencode(".jpg", face_crop)
        b = buf.tobytes()
        key = s3_key(PROFILES_PREFIX, uid, f"{uid}_img{next_img_num}.jpg")
        s3_put_bytes(key, b, "image/jpeg")
        relative_path = key
        existing_record.setdefault("profile_img", []).append(relative_path)
        put_face_item(uid, existing_record)
        return jsonify(
            {
                "success": True,
                "message": "Profile image processed successfully",
                "image_path": relative_path,
                "image_number": next_img_num,
            }
        )
    except Exception as e:
        return jsonify({"success": False, "message": f"Error processing image: {str(e)}"}), 500
    finally:
        cleanup_memory()


@app.route("/departments", methods=["GET"])
def get_departments():
    table = dynamodb.Table(DEPT_TABLE)
    try:
        resp = table.scan()
        items = resp.get("Items", [])
        while "LastEvaluatedKey" in resp:
            resp = table.scan(ExclusiveStartKey=resp["LastEvaluatedKey"])
            items.extend(resp.get("Items", []))
        return jsonify({"departments": items})
    except Exception:
        return jsonify({"error": "Departments not found"}), 404


@app.route("/complete_meeting", methods=["POST"])
def complete_meeting():
    try:
        visit_id = request.args.get("visit_id")
        uid = request.args.get("uid")
        if not visit_id or not uid:
            return jsonify({"error": "Missing visit_id or uid parameter"}), 400
        record = get_face_item(uid)
        if not record:
            return jsonify({"error": f"User with UID {uid} not found"}), 404
        visit_found = False
        for visit in record.get("visitor", []):
            if visit.get("visit_id") == visit_id:
                visit["status"] = "completed"
                visit_found = True
                break
        if not visit_found:
            return jsonify({"error": f"Visit with ID {visit_id} not found for user {uid}"}), 404
        put_face_item(uid, record)
        return jsonify({"success": True, "message": f"Meeting {visit_id} marked as completed for user {uid}"}), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/visitor_action", methods=["POST"])
def visitor_action():
    if "user" not in session:
        return jsonify({"success": False, "message": "Unauthorized"}), 401
    data = request.get_json()
    if not data:
        return jsonify({"success": False, "message": "Invalid JSON payload."}), 400
    visit_id = data.get("visit_id")
    uid = data.get("uid")
    remark = data.get("remark", "")
    forwarding_department = data.get("forwarding_department")
    if not visit_id or not uid or not forwarding_department:
        missing = []
        if not visit_id:
            missing.append("visit_id")
        if not uid:
            missing.append("uid")
        if not forwarding_department:
            missing.append("forwarding_department")
        return jsonify({"success": False, "message": f"Missing required fields: {', '.join(missing)}"}), 400
    record = get_face_item(uid)
    if not record:
        return jsonify({"success": False, "message": f"Visitor with UID '{uid}' not found."}), 404
    visit_found = False
    for visit in record.get("visitor", []):
        if visit.get("visit_id") == visit_id:
            visit["remark"] = remark
            visit["forwarding_department"] = forwarding_department
            visit["status"] = "pending"
            visit_found = True
            break
    if not visit_found:
        return jsonify({"success": False, "message": f"Visit with ID '{visit_id}' not found for visitor '{uid}'."}), 404
    put_face_item(uid, record)
    return jsonify({"success": True, "message": "Visitor action recorded successfully."}), 200


@app.route("/registered_visitor_today")
def registered_visitor_today():
    data = scan_face_data()
    today = datetime.now().strftime("%Y-%m-%d")
    today_visitors = []
    for uid, user_data in data.items():
        if "visitor" in user_data:
            for visit in user_data["visitor"]:
                if "datetime" in visit and isinstance(visit["datetime"], str):
                    visit_date_str = visit["datetime"].split(" ")[0]
                    try:
                        datetime.strptime(visit_date_str, "%Y-%m-%d")
                        if visit_date_str == today:
                            today_visitors.append(
                                {
                                    "uid": uid,
                                    "name": user_data.get("name", "N/A"),
                                    "phone": user_data.get("phone", "N/A"),
                                    "purpose": visit.get("purpose", "N/A"),
                                    "visit_id": visit.get("visit_id", "N/A"),
                                    "status": visit.get("status", "Unknown"),
                                    "pdf_path": visit.get("document_pdf", ""),
                                }
                            )
                    except ValueError:
                        print(
                            f"Warning: Invalid date format found for visit_id {visit.get('visit_id', 'N/A')}: {visit['datetime']}"
                        )
                else:
                    print(f"Warning: Missing or invalid 'datetime' for visit in user {uid}")
    return jsonify({"visitors": today_visitors})


@app.route("/api/visitor_file_status")
def get_visitor_file_status():
    try:
        today = datetime.today().strftime("%Y-%m-%d")
        data = scan_face_data()
        visitors = []
        for uid, u in data.items():
            if not isinstance(u, dict):
                continue
            dt = u.get("registration_datetime", "")
            try:
                if datetime.fromisoformat(dt).strftime("%Y-%m-%d") != today:
                    continue
            except Exception:
                continue
            enc = any(s3_exists(p) for p in u.get("face_encodings", []) if isinstance(p, str))
            img = any(s3_exists(p) for p in u.get("profile_img", []) if isinstance(p, str))
            visitors.append({"uid": uid, "name": u.get("name", "N/A"), "has_encodings": enc, "has_profile_img": img})
        return jsonify({"success": True, "visitors": sorted(visitors, key=lambda x: x["name"])})
    except Exception as e:
        traceback.print_exc()
        return jsonify({"success": False, "message": f"Server error: {str(e)}"}), 500


@app.route("/api/upload-visitor-document/<visit_id>", methods=["POST"])
def upload_visitor_document(visit_id):
    if "document" not in request.files:
        return jsonify({"success": False, "error": "No document file part in the request"}), 400
    file = request.files["document"]
    if file.filename == "":
        return jsonify({"success": False, "error": "No file selected"}), 400
    if not file or not allowed_file(file.filename):
        return jsonify({"success": False, "error": "Invalid file type. Only PDF allowed."}), 400
    try:
        data = scan_face_data()
        found_uid = None
        found_visit = None
        for uid, user_data in data.items():
            for visit in user_data.get("visitor", []):
                if visit.get("visit_id") == visit_id:
                    found_uid = uid
                    found_visit = visit
                    break
            if found_visit:
                break
        if not found_visit:
            return jsonify({"success": False, "error": "Visitor ID not found"}), 404
        existing_pdf = found_visit.get("document_pdf")
        if existing_pdf and isinstance(existing_pdf, str) and s3_exists(existing_pdf):
            s3_delete(existing_pdf)
        filename = f"{secure_filename(visit_id)}_doc.pdf"
        key = s3_key(PDF_PREFIX, filename)
        b = file.read()
        s3_put_bytes(key, b, "application/pdf")
        found_visit["document_pdf"] = key
        record = get_face_item(found_uid)
        for v in record.get("visitor", []):
            if v.get("visit_id") == visit_id:
                v["document_pdf"] = key
                break
        put_face_item(found_uid, record)
        return jsonify({"success": True, "message": "Document uploaded successfully", "pdf_path": key})
    except Exception as e:
        traceback.print_exc()
        return jsonify({"success": False, "error": f"An internal error occurred: {str(e)}"}), 500


@app.route("/api/delete-visitor-document/<visit_id>", methods=["POST"])
def delete_visitor_document(visit_id):
    try:
        data = scan_face_data()
        found_uid = None
        found_visit = None
        for uid, user_data in data.items():
            for visit in user_data.get("visitor", []):
                if visit.get("visit_id") == visit_id:
                    found_uid = uid
                    found_visit = visit
                    break
            if found_visit:
                break
        if not found_visit:
            return jsonify({"success": False, "message": "Visitor ID not found"}), 404
        pdf_path = found_visit.get("document_pdf")
        if not pdf_path or not isinstance(pdf_path, str):
            return jsonify({"success": False, "message": "No document found associated with this visit entry"}), 404
        file_deleted = False
        if s3_exists(pdf_path):
            try:
                s3_delete(pdf_path)
                file_deleted = True
            except OSError as e:
                return jsonify({"success": False, "message": f"Error deleting file: {e}"}), 500
        else:
            file_deleted = True
        if file_deleted:
            found_visit["document_pdf"] = None
            record = get_face_item(found_uid)
            for v in record.get("visitor", []):
                if v.get("visit_id") == visit_id:
                    v["document_pdf"] = None
                    break
            put_face_item(found_uid, record)
            return jsonify(
                {"success": True, "message": "Document reference removed and file deleted successfully (if it existed)."}
            )
        else:
            return jsonify({"success": False, "message": "File deletion was prevented or failed."}), 500
    except Exception as e:
        traceback.print_exc()
        return jsonify({"success": False, "message": f"An internal server error occurred: {str(e)}"}), 500


@app.route("/api/get-visitor-document/<visit_id>")
def get_visitor_document(visit_id):
    data = scan_face_data()
    pdf_key = None
    for uid, user_data in data.items():
        for visit in user_data.get("visitor", []):
            if visit.get("visit_id") == visit_id and "document_pdf" in visit:
                full_pdf_path = visit["document_pdf"]
                if full_pdf_path and isinstance(full_pdf_path, str):
                    pdf_key = full_pdf_path
                    break
        if pdf_key:
            break
    if pdf_key:
        try:
            b = s3_get_bytes(pdf_key)
            return send_file(io.BytesIO(b), download_name=os.path.basename(pdf_key), mimetype="application/pdf")
        except FileNotFoundError:
            return jsonify({"error": "Document file not found on server."}), 404
        except Exception:
            return jsonify({"error": "An error occurred while retrieving the document."}), 500
    else:
        return jsonify({"error": "Document reference not found for this visitor ID."}), 404


@app.route("/api/upload-image", methods=["POST"])
def upload_image():
    if "image" not in request.json:
        return jsonify({"error": "No image provided"}), 400
    try:
        image_data = request.json["image"].split(",")[1]
        image_bytes = base64.b64decode(image_data)
        filename = f"{uuid.uuid4()}.jpg"
        key = s3_key(UPLOAD_PREFIX, filename)
        s3_put_bytes(key, image_bytes, "image/jpeg")
        return jsonify({"success": True, "image_id": key})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/generate-pdf", methods=["POST"])
def generate_pdf():
    if "images" not in request.json or "visit_id" not in request.json:
        return jsonify({"success": False, "error": "No images or visit ID provided"}), 400
    try:
        image_ids = request.json["images"]
        visit_id = request.json["visit_id"]
        if not visit_id:
            return jsonify({"success": False, "error": "Visit ID is missing"}), 400
        if not image_ids:
            return jsonify({"success": False, "error": "No image IDs provided for PDF generation"}), 400
        pdf_filename = f"{secure_filename(visit_id)}_doc.pdf"
        pdf_key = s3_key(PDF_PREFIX, pdf_filename)
        buffer = io.BytesIO()
        c = canvas.Canvas(buffer, pagesize=letter)
        letter_width, letter_height = letter
        for image_key in image_ids:
            try:
                b = s3_get_bytes(image_key)
                img = Image.open(io.BytesIO(b))
                width, height = img.size
                scale = min(letter_width / width, letter_height / height) * 0.9
                img_width = width * scale
                img_height = height * scale
                x = (letter_width - img_width) / 2
                y = (letter_height - img_height) / 2
                tmp_path = f"/tmp/{uuid.uuid4()}.jpg"
                img.save(tmp_path, "JPEG")
                c.drawImage(tmp_path, x, y, width=img_width, height=img_height, preserveAspectRatio=True, mask="auto")
                c.showPage()
                try:
                    os.remove(tmp_path)
                except Exception:
                    pass
            except Exception:
                pass
        c.save()
        buffer.seek(0)
        s3_put_bytes(pdf_key, buffer.read(), "application/pdf")
        buffer.close()
        data = scan_face_data()
        updated = False
        for uid, user_data in data.items():
            for visit in user_data.get("visitor", []):
                if visit.get("visit_id") == visit_id:
                    visit["document_pdf"] = pdf_key
                    put_face_item(uid, user_data)
                    updated = True
                    break
            if updated:
                break
        return jsonify({"success": True, "message": "PDF generated successfully from scanned images.", "pdf_url": f"/api/get-visitor-document/{visit_id}"})
    except Exception as e:
        traceback.print_exc()
        return jsonify({"success": False, "error": f"Failed to generate PDF: {str(e)}"}), 500


@app.route("/api/cleanup", methods=["POST"])
def cleanup_images():
    try:
        resp = s3.list_objects_v2(Bucket=BUCKET, Prefix=s3_key(UPLOAD_PREFIX))
        for obj in resp.get("Contents", []):
            s3_delete(obj["Key"])
        return jsonify({"success": True, "message": "All temporary images have been deleted"})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


def load_departments():
    table = dynamodb.Table(DEPT_TABLE)
    try:
        resp = table.scan()
        items = resp.get("Items", [])
        valid_departments = [d for d in items if d.get("name") and d.get("email")]
        return valid_departments
    except Exception:
        return []


def load_departments_data():
    table = dynamodb.Table(DEPT_TABLE)
    try:
        resp = table.scan()
        items = resp.get("Items", [])
        return {"departments": items}
    except Exception:
        return {"departments": []}


def save_departments_data(data):
    table = dynamodb.Table(DEPT_TABLE)
    try:
        for d in data.get("departments", []):
            table.put_item(Item=d)
        return True
    except Exception:
        return False


def send_otp_email(recipient_email, otp):
    import smtplib
    import ssl
    import random
    from email.mime.multipart import MIMEMultipart
    from email.mime.text import MIMEText
    from email.utils import formataddr, formatdate

    sender_email = "svmsakola@gmail.com"
    sender_password = os.environ.get("SVMS_SMTP_PASS", "aess hmcl bkqm slph")
    sender_name = "स्मार्ट व्हिजिटर मॅनेजमेंट सिस्टिम (SVMS)"
    message = MIMEMultipart("alternative")
    message["Subject"] = "सुरक्षित प्रवेश कोड - SVMS विभाग पोर्टल"
    message["From"] = formataddr((sender_name, sender_email))
    message["To"] = recipient_email
    message["Date"] = formatdate(localtime=True)
    message["Message-ID"] = f"<{random.getrandbits(128)}@{sender_email.split('@')[-1]}>"
    message["X-Mailer"] = "SVMSMailer/1.0"
    message["Reply-To"] = sender_email
    message["Return-Path"] = sender_email

    text = f"""आदरणीय SVMS विभाग सदस्य,

आपला सुरक्षित प्रवेश कोड:

{otp}

हा कोड 10 मिनिटांत कालबाह्य होईल.

महत्वाचे: कृपया हा कोड कोणाशीही शेअर करू नका.

जर आपण हा विनंती केलेली नसेल तर कृपया आमच्या IT टीमशी तात्काळ संपर्क साधा.

सादर,
स्मार्ट व्हिजिटर मॅनेजमेंट सिस्टिम (SVMS)
"""

    html = (
        "<!DOCTYPE html><html lang=\"mr\"><head><meta charset=\"UTF-8\"><title>SVMS Secure Code</title></head>"
        "<body style=\"margin:0;padding:0;font-family:'Helvetica Neue',Helvetica,Arial,sans-serif;background-color:#f4f6f9;\">"
        "<table align=\"center\" width=\"600\" cellpadding=\"0\" cellspacing=\"0\" style=\"background:#ffffff;border-radius:8px;overflow:hidden;box-shadow:0 2px 8px rgba(0,0,0,0.1);\">"
        "<tr><td align=\"center\" bgcolor=\"#004aad\" style=\"padding:40px 0;\"><h1 style=\"color:#ffffff;font-size:28px;margin:0;\">स्मार्ट व्हिजिटर मॅनेजमेंट सिस्टिम (SVMS)</h1></td></tr>"
        "<tr><td style=\"padding:30px 40px;\"><h2 style=\"color:#333333;font-size:22px;margin:0 0 20px;\">आपला सुरक्षित प्रवेश कोड</h2>"
        "<p style=\"color:#555555;font-size:16px;margin:0 0 10px;\">आदरणीय SVMS विभाग सदस्य,</p>"
        "<p style=\"color:#555555;font-size:16px;margin:10px 0;\">आपला प्रवेश कोड खाली दिला आहे:</p>"
        f"<div style=\"background:#eef4fc;border:1px dashed #004aad;border-radius:6px;text-align:center;padding:20px;margin:20px 0;font-size:28px;font-weight:bold;letter-spacing:5px;color:#004aad;\">{otp}</div>"
        "<p style=\"color:#777777;font-size:14px;margin:20px 0;\">हा कोड <strong>10 मिनिटांत</strong> कालबाह्य होईल. कृपया सुरक्षा कारणास्तव कोड कोणाशीही शेअर करू नका.</p>"
        "<p style=\"color:#e74c3c;font-weight:bold;font-size:14px;\">महत्वाचे: SVMS स्टाफ कधीही आपल्याकडून हा कोड मागणार नाही.</p>"
        "<p style=\"color:#555555;font-size:14px;margin:20px 0;\">जर आपण हा विनंती केलेली नसेल तर त्वरित आमच्या IT सहाय्यता टीमशी संपर्क साधा.</p>"
        "</td></tr>"
        f"<tr><td bgcolor=\"#f0f0f0\" style=\"padding:20px;text-align:center;color:#999999;font-size:12px;\"><p style=\"margin:0;\">© {sender_name} सर्व हक्क राखीव.</p><p style=\"margin:5px 0 0;\">ही एक स्वयंचलित ई-मेल आहे, कृपया प्रतिसाद देऊ नका.</p></td></tr>"
        "</table></body></html>"
    )

    message.attach(MIMEText(text, "plain"))
    message.attach(MIMEText(html, "html"))
    try:
        context = ssl.create_default_context()
        context.check_hostname = False
        context.verify_mode = ssl.CERT_NONE
        with smtplib.SMTP_SSL("smtp.gmail.com", 465, context=context) as server:
            server.login(sender_email, sender_password)
            server.sendmail(sender_email, recipient_email, message.as_string())
            print(f"Access code sent to {recipient_email}")
            return True
    except Exception as e:
        print(f"Email error: {str(e)}")
        return False


def admin_login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if "user" not in session or session.get("user", {}).get("role") != "admin":
            return jsonify({"success": False, "message": "Admin access required."}), 403
        return f(*args, **kwargs)

    return decorated_function


def department_login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if "department" not in session:
            flash("Please login to access this page.", "warning")
            return redirect(url_for("login_department"))
        return f(*args, **kwargs)

    return decorated_function


@app.route("/department/login", methods=["GET", "POST"])
def login_department():
    if "department" in session:
        return redirect(url_for("department_dashboard"))
    departments_list = load_departments()
    if request.method == "POST":
        entered_email = request.form.get("department_email")
        if not entered_email:
            flash("Please enter the department email address.", "warning")
            return render_template("department/auth/login_department.html")
        department = next((d for d in departments_list if d.get("email") == entered_email), None)
        if department:
            department_name = department.get("name", "Unknown Department")
            email = department.get("email")
            otp = str(random.randint(100000, 999999))
            session["otp"] = otp
            session["otp_timestamp"] = datetime.now().timestamp()
            session["pending_department_info"] = department
            if send_otp_email(email, otp):
                flash(f"An OTP has been sent to {email} for {department_name}. Please check the inbox.", "info")
                return redirect(url_for("verify_otp"))
            else:
                session.pop("otp", None)
                session.pop("otp_timestamp", None)
                session.pop("pending_department_info", None)
                flash("Failed to send OTP email. Please check server logs or contact support.", "danger")
                return render_template("department/auth/login_department.html")
        else:
            flash("Email address not registered for any department.", "danger")
            return render_template("department/auth/login_department.html")
    return render_template("department/auth/login_department.html")


@app.route("/department/verify-otp", methods=["GET", "POST"])
def verify_otp():
    pending_info = session.get("pending_department_info")
    display_email = pending_info.get("email", "your registered email") if pending_info else "your registered email"
    department_name = pending_info.get("name", "Your Department") if pending_info else "Your Department"
    if request.method == "POST":
        entered_otp = request.form.get("otp")
        stored_otp = session.get("otp")
        otp_timestamp = session.get("otp_timestamp")
        if not entered_otp:
            flash("Please enter the OTP.", "warning")
        elif not stored_otp or not otp_timestamp:
            flash("OTP session expired or invalid. Please login again.", "danger")
            session.pop("otp", None)
            session.pop("otp_timestamp", None)
            session.pop("pending_department_info", None)
            return redirect(url_for("login_department"))
        elif (datetime.now().timestamp() - otp_timestamp) > 300:
            flash("OTP has expired (valid for 5 minutes). Please request a new one.", "danger")
            session.pop("otp", None)
            session.pop("otp_timestamp", None)
            session.pop("pending_department_info", None)
            return redirect(url_for("login_department"))
        elif entered_otp == stored_otp:
            session["department"] = session.pop("pending_department_info", None)
            session.pop("otp", None)
            session.pop("otp_timestamp", None)
            session.modified = True
            if not session.get("department"):
                flash("Login failed due to session issue. Please try again.", "danger")
                return redirect(url_for("login_department"))
            flash("Login successful!", "success")
            return redirect(url_for("department_dashboard"))
        else:
            flash("Invalid OTP entered. Please try again.", "danger")
    return render_template("department/auth/verify_otp.html", email=display_email, department_name=department_name)


@app.route("/department/dashboard")
@department_login_required
def department_dashboard():
    return render_template("department/dashboard/department_dashboard.html", department=session.get("department"))


@app.route("/department/logout")
def logout_department():
    session.pop("department", None)
    session.pop("otp", None)
    session.pop("otp_timestamp", None)
    session.pop("pending_department_info", None)
    flash("You have been logged out successfully.", "info")
    return redirect(url_for("login_department"))


@app.route("/api/department/applications", methods=["GET"])
@department_login_required
def get_department_applications():
    department = session.get("department")
    department_name = department.get("name")
    start_date = request.args.get("start_date")
    end_date = request.args.get("end_date")
    status = request.args.get("status")
    facedata = scan_face_data()
    applications = []
    for uid, user_data in facedata.items():
        for visit in user_data.get("visitor", []):
            if visit.get("forwarding_department") == department_name:
                if status and visit.get("status") != status:
                    continue
                if start_date or end_date:
                    try:
                        visit_date = datetime.strptime(visit.get("datetime"), "%Y-%m-%d %H:%M:%S").date()
                        if start_date:
                            start = datetime.strptime(start_date, "%Y-%m-%d").date()
                            if visit_date < start:
                                continue
                        if end_date:
                            end = datetime.strptime(end_date, "%Y-%m-%d").date()
                            if visit_date > end:
                                continue
                    except ValueError:
                        continue
                application = {
                    "uid": uid,
                    "visit_id": visit.get("visit_id"),
                    "name": user_data.get("name"),
                    "phone": user_data.get("phone"),
                    "email": user_data.get("email"),
                    "datetime": visit.get("datetime"),
                    "purpose": visit.get("purpose"),
                    "status": visit.get("status"),
                    "remark": visit.get("remark", ""),
                    "profile_img": user_data.get("profile_img", [""])[0] if user_data.get("profile_img") else "",
                }
                applications.append(application)
    return jsonify(applications)


@app.route("/data/remarkpdf/<path:filename>")
@department_login_required
def serve_remark_pdf(filename):
    key = s3_key(REMARK_PDF_PREFIX, filename)
    try:
        b = s3_get_bytes(key)
        return send_file(io.BytesIO(b), download_name=filename, mimetype="application/pdf")
    except Exception:
        return jsonify({"error": "File not found"}), 404


@app.route("/api/department/application/<visit_id>", methods=["GET"])
@department_login_required
def get_application_details(visit_id):
    department = session.get("department")
    department_name = department.get("name")
    data = scan_face_data()
    for uid, user_data in data.items():
        for visit in user_data.get("visitor", []):
            if visit.get("visit_id") == visit_id:
                if visit.get("forwarding_department") == department_name:
                    application = {
                        "uid": uid,
                        "visit_id": visit.get("visit_id"),
                        "name": user_data.get("name"),
                        "phone": user_data.get("phone"),
                        "email": user_data.get("email"),
                        "address": user_data.get("address"),
                        "district": user_data.get("district"),
                        "tahasil": user_data.get("tahasil"),
                        "registration_datetime": user_data.get("registration_datetime"),
                        "datetime": visit.get("datetime"),
                        "purpose": visit.get("purpose"),
                        "status": visit.get("status"),
                        "remark": visit.get("remark", ""),
                        "remark_pdf": visit.get("remark_pdf", ""),
                        "document_pdf": visit.get("document_pdf", ""),
                        "entry_confirmed": visit.get("entry_confirmed", False),
                        "confirmation_time": visit.get("confirmation_time", ""),
                        "profile_img": user_data.get("profile_img", [""])[0] if user_data.get("profile_img") else "",
                        "images": user_data.get("images", []),
                    }
                    return jsonify(application)
                else:
                    return jsonify({"error": "Unauthorized access"}), 403
    return jsonify({"error": "Application not found"}), 404


@app.route("/api/department/update-application/<visit_id>", methods=["POST"])
@department_login_required
def update_application(visit_id):
    department = session.get("department")
    department_name = department.get("name")
    status = request.form.get("status")
    remark = request.form.get("remark", "")
    if not status or status not in ["pending", "completed"]:
        return jsonify({"error": "Invalid status value"}), 400
    data = scan_face_data()
    updated = False
    for uid, user_data in data.items():
        for i, visit in enumerate(user_data.get("visitor", [])):
            if visit.get("visit_id") == visit_id:
                if visit.get("forwarding_department") == department_name:
                    user_data["visitor"][i]["status"] = status
                    user_data["visitor"][i]["remark"] = remark
                    if "remark_pdf" in request.files:
                        file = request.files["remark_pdf"]
                        if file and file.filename != "" and allowed_file(file.filename):
                            old_pdf_path = user_data["visitor"][i].get("remark_pdf")
                            if old_pdf_path and s3_exists(old_pdf_path):
                                s3_delete(old_pdf_path)
                            filename = f"{secure_filename(visit_id)}_remark_{int(datetime.now().timestamp())}.pdf"
                            key = s3_key(REMARK_PDF_PREFIX, filename)
                            s3_put_bytes(key, file.read(), "application/pdf")
                            user_data["visitor"][i]["remark_pdf"] = key
                        elif file.filename != "":
                            return jsonify({"error": "Invalid file type. Only PDF is allowed."}), 400
                    put_face_item(uid, user_data)
                    updated = True
                    break
        if updated:
            break
    if updated:
        return jsonify({"success": True, "message": "Application updated successfully"})
    else:
        return jsonify({"error": "Application not found or unauthorized"}), 404


@app.route("/api/department/application-stats", methods=["GET"])
@department_login_required
def get_application_stats():
    department = session.get("department")
    department_name = department.get("name")
    data = scan_face_data()
    total_count = 0
    pending_count = 0
    completed_count = 0
    today = datetime.now().date()
    today_count = 0
    for uid, user_data in data.items():
        for visit in user_data.get("visitor", []):
            if visit.get("forwarding_department") == department_name:
                total_count += 1
                if visit.get("status") == "pending":
                    pending_count += 1
                elif visit.get("status") == "completed":
                    completed_count += 1
                try:
                    visit_date = datetime.strptime(visit.get("datetime"), "%Y-%m-%d %H:%M:%S").date()
                    if visit_date == today:
                        today_count += 1
                except ValueError:
                    pass
    stats = {"total": total_count, "pending": pending_count, "completed": completed_count, "today": today_count}
    return jsonify(stats)


@app.route("/api/department/forward-application/<visit_id>", methods=["POST"])
@department_login_required
def forward_application(visit_id):
    current_department = session.get("department")
    current_department_name = current_department.get("name")
    target_department = request.form.get("target_department")
    note = request.form.get("note", "")
    if not target_department:
        return jsonify({"error": "Target department is required"}), 400
    departments = load_departments()
    if not any(d["name"] == target_department for d in departments):
        return jsonify({"error": "Invalid target department"}), 400
    data = scan_face_data()
    updated = False
    for uid, user_data in data.items():
        for i, visit in enumerate(user_data.get("visitor", [])):
            if visit.get("visit_id") == visit_id:
                if visit.get("forwarding_department") == current_department_name:
                    user_data["visitor"][i]["forwarding_department"] = target_department
                    user_data["visitor"][i]["status"] = "pending"
                    user_data["visitor"][i]["forwarding_note"] = note
                    user_data["visitor"][i]["forwarded_from"] = current_department_name
                    user_data["visitor"][i]["forwarded_datetime"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    put_face_item(uid, user_data)
                    updated = True
                    break
        if updated:
            break
    if updated:
        return jsonify({"success": True, "message": f"Application forwarded to {target_department} successfully"})
    else:
        return jsonify({"error": "Application not found or unauthorized"}), 404


@app.route("/api/department/search-applications", methods=["GET"])
@department_login_required
def search_applications():
    department = session.get("department")
    department_name = department.get("name")
    query = request.args.get("query", "").lower()
    if not query or len(query) < 3:
        return jsonify({"error": "Search query must be at least 3 characters"}), 400
    data = scan_face_data()
    results = []
    for uid, user_data in data.items():
        for visit in user_data.get("visitor", []):
            if visit.get("forwarding_department") == department_name:
                if (
                    query in user_data.get("name", "").lower()
                    or query in user_data.get("phone", "").lower()
                    or query in user_data.get("email", "").lower()
                    or query in visit.get("visit_id", "").lower()
                    or query in visit.get("purpose", "").lower()
                    or query in user_data.get("address", "").lower()
                    or query in user_data.get("district", "").lower()
                ):
                    result = {
                        "uid": uid,
                        "visit_id": visit.get("visit_id"),
                        "name": user_data.get("name"),
                        "phone": user_data.get("phone"),
                        "email": user_data.get("email"),
                        "datetime": visit.get("datetime"),
                        "purpose": visit.get("purpose"),
                        "status": visit.get("status"),
                        "profile_img": user_data.get("profile_img", [""])[0] if user_data.get("profile_img") else "",
                    }
                    results.append(result)
    return jsonify(results)


@app.route("/api/department/recent-activities", methods=["GET"])
@department_login_required
def get_recent_activities():
    department = session.get("department")
    department_name = department.get("name")
    limit = int(request.args.get("limit", 10))
    data = scan_face_data()
    activities = []
    for uid, user_data in data.items():
        for visit in user_data.get("visitor", []):
            if visit.get("forwarding_department") == department_name:
                activity = {
                    "uid": uid,
                    "visit_id": visit.get("visit_id"),
                    "name": user_data.get("name"),
                    "datetime": visit.get("datetime"),
                    "purpose": visit.get("purpose"),
                    "status": visit.get("status"),
                    "profile_img": user_data.get("profile_img", [""])[0] if user_data.get("profile_img") else "",
                }
                activities.append(activity)
    try:
        activities.sort(key=lambda x: datetime.strptime(x["datetime"], "%Y-%m-%d %H:%M:%S"), reverse=True)
    except Exception:
        pass
    return jsonify(activities[:limit])


@app.route("/api/admin/departments", methods=["GET"])
@admin_login_required
def admin_get_departments():
    return jsonify(load_departments_data().get("departments", []))


@app.route("/api/admin/departments", methods=["POST"])
@admin_login_required
def admin_add_department():
    req_data = request.get_json()
    if not req_data or not req_data.get("name") or not req_data.get("email"):
        return jsonify({"success": False, "message": "Name and email are required."}), 400
    name = req_data["name"].strip()
    email = req_data["email"].strip()
    data = load_departments_data()
    departments = data.get("departments", [])
    if any(d["name"].lower() == name.lower() for d in departments):
        return jsonify({"success": False, "message": "A department with this name already exists."}), 409
    if any(d["email"].lower() == email.lower() for d in departments):
        return jsonify({"success": False, "message": "A department with this email already exists."}), 409
    new_id = max([d.get("id", 0) for d in departments]) + 1 if departments else 1
    new_department = {"id": new_id, "name": name, "email": email}
    table = dynamodb.Table(DEPT_TABLE)
    table.put_item(Item=new_department)
    return jsonify({"success": True, "message": "Department added successfully.", "department": new_department}), 201


@app.route("/api/admin/departments/<int:dept_id>", methods=["PUT"])
@admin_login_required
def admin_update_department(dept_id):
    req_data = request.get_json()
    if not req_data:
        return jsonify({"success": False, "message": "Request body is empty."}), 400
    new_name = req_data.get("name", "").strip()
    new_email = req_data.get("email", "").strip()
    if not new_name and not new_email:
        return jsonify({"success": False, "message": "At least one field (name or email) must be provided for update."}), 400
    data = load_departments_data()
    departments = data.get("departments", [])
    target_dept = next((d for d in departments if d.get("id") == dept_id), None)
    if not target_dept:
        return jsonify({"success": False, "message": "Department not found."}), 404
    if new_name and any(d["name"].lower() == new_name.lower() and d["id"] != dept_id for d in departments):
        return jsonify({"success": False, "message": "Another department with this name already exists."}), 409
    if new_email and any(d["email"].lower() == new_email.lower() and d["id"] != dept_id for d in departments):
        return jsonify({"success": False, "message": "Another department with this email already exists."}), 409
    if new_name:
        target_dept["name"] = new_name
    if new_email:
        target_dept["email"] = new_email
    save_departments_data({"departments": departments})
    return jsonify({"success": True, "message": "Department updated successfully.", "department": target_dept})


@app.route("/api/admin/departments/<int:dept_id>", methods=["DELETE"])
@admin_login_required
def admin_delete_department(dept_id):
    data = load_departments_data()
    departments = data.get("departments", [])
    initial_len = len(departments)
    departments = [d for d in departments if d.get("id") != dept_id]
    if len(departments) == initial_len:
        return jsonify({"success": False, "message": "Department not found."}), 404
    save_departments_data({"departments": departments})
    return jsonify({"success": True, "message": "Department deleted successfully."})


def admin_or_inoffice_login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if "user" not in session:
            return jsonify({"success": False, "message": "Authentication required."}), 401
        user_role = session.get("user", {}).get("role")
        if user_role not in ["admin", "inoffice"]:
            return jsonify({"success": False, "message": "Admin or In-Office access required."}), 403
        return f(*args, **kwargs)

    return decorated_function


@app.route("/api/admin/visitors")
@admin_or_inoffice_login_required
def get_admin_visitors():
    selected_date_str = request.args.get("date")
    if not selected_date_str:
        return jsonify({"error": "Date parameter is required"}), 400
    try:
        datetime.strptime(selected_date_str, "%Y-%m-%d")
    except ValueError:
        return jsonify({"error": "Invalid date format. Use YYYY-MM-DD"}), 400
    face_data = scan_face_data()
    daily_visitors = []
    for uid, user_data in face_data.items():
        if isinstance(user_data, dict) and "visitor" in user_data and isinstance(user_data["visitor"], list):
            for visit in user_data["visitor"]:
                if isinstance(visit, dict) and "datetime" in visit and isinstance(visit["datetime"], str):
                    try:
                        if visit["datetime"].split(" ")[0] == selected_date_str:
                            daily_visitors.append(
                                {
                                    "uid": uid,
                                    "name": user_data.get("name", "N/A"),
                                    "phone": user_data.get("phone", "N/A"),
                                    "address": user_data.get("address", "N/A"),
                                    "tahasil": user_data.get("tahasil", "N/A"),
                                    "district": user_data.get("district", "N/A"),
                                    "profile_img": user_data.get("profile_img", user_data.get("images", [])[:1]),
                                    "images": user_data.get("images", []),
                                    "face_encodings": user_data.get("face_encodings", []),
                                    "visit_id": visit.get("visit_id", "N/A"),
                                    "datetime": visit.get("datetime", ""),
                                    "purpose": visit.get("purpose", "N/A"),
                                    "status": visit.get("status", "unknown"),
                                    "entry_confirmed": visit.get("entry_confirmed", False),
                                    "dvn": visit.get("dvn", None),
                                    "forwarding_department": visit.get("forwarding_department", ""),
                                    "forwarding_note": visit.get("forwarding_note", ""),
                                    "remark": visit.get("remark", ""),
                                    "document_pdf": visit.get("document_pdf", None),
                                }
                            )
                    except Exception as e:
                        print(f"Error processing visit for UID {uid}, Visit {visit.get('visit_id', 'N/A')}: {e}")
                else:
                    print(f"Warning: Skipping invalid visit entry for UID {uid}: {visit}")
        else:
            print(f"Warning: Skipping invalid user data structure for UID {uid}")
    try:
        valid_visitors = [v for v in daily_visitors if "datetime" in v and isinstance(v["datetime"], str) and v["datetime"]]
        invalid_visitors = [v for v in daily_visitors if not ("datetime" in v and isinstance(v["datetime"], str) and v["datetime"])]
        valid_visitors.sort(key=lambda x: datetime.strptime(x["datetime"], "%Y-%m-%d %H:%M:%S"))
        sorted_daily_visitors = valid_visitors + invalid_visitors
    except Exception:
        sorted_daily_visitors = daily_visitors
    response = make_response(jsonify({"visitors": sorted_daily_visitors}))
    response.headers["Cache-Control"] = "no-cache, no-store, must-revalidate"
    response.headers["Pragma"] = "no-cache"
    response.headers["Expires"] = "0"
    return response


@app.route("/api/admin/delete_visitor_entry", methods=["POST"])
@admin_login_required
def delete_visitor_entry():
    data = request.get_json()
    uid = data.get("uid") if data else None
    visit_id = data.get("visit_id") if data else None
    if not uid or not visit_id:
        return jsonify({"success": False, "message": "Missing or empty uid/visit_id"}), 400
    record = get_face_item(uid)
    if not record:
        return jsonify({"success": False, "message": "Invalid or missing visitor data"}), 404
    visit_to_remove = next((v for v in record.get("visitor", []) if isinstance(v, dict) and v.get("visit_id") == visit_id), None)
    if not visit_to_remove:
        return jsonify({"success": False, "message": f"Visit ID '{visit_id}' not found for UID '{uid}'"}), 404
    record["visitor"] = [v for v in record.get("visitor", []) if v.get("visit_id") != visit_id]
    pdf = visit_to_remove.get("document_pdf")
    if pdf and isinstance(pdf, str) and pdf.strip() and s3_exists(pdf):
        s3_delete(pdf)
    put_face_item(uid, record)
    return jsonify({"success": True, "message": "Visitor entry deleted successfully"})


def safe_get(data, key, default="N/A"):
    return data.get(key, default) if data else default


@app.route("/api/admin/generate_report")
@admin_login_required
def generate_visitor_report():
    start_date_str = request.args.get("start_date")
    end_date_str = request.args.get("end_date")
    if not start_date_str:
        return jsonify({"error": "Missing required parameter: start_date"}), 400
    try:
        start_date = datetime.strptime(start_date_str, "%Y-%m-%d").date()
        end_date = start_date if not end_date_str else datetime.strptime(end_date_str, "%Y-%m-%d").date()
        if start_date > end_date:
            return jsonify({"error": "Start date cannot be after end date"}), 400
    except ValueError:
        return jsonify({"error": "Invalid date format. Use YYYY-MM-DD"}), 400
    try:
        face_data = scan_face_data()
        report_data = []
        for uid, user_data in face_data.items():
            if isinstance(user_data, dict) and "visitor" in user_data and isinstance(user_data["visitor"], list):
                for visit in user_data["visitor"]:
                    if isinstance(visit, dict) and "datetime" in visit and isinstance(visit["datetime"], str):
                        try:
                            visit_dt = datetime.strptime(visit["datetime"], "%Y-%m-%d %H:%M:%S")
                            visit_date = visit_dt.date()
                            if start_date <= visit_date <= end_date:
                                report_entry = {
                                    "Name": safe_get(user_data, "name"),
                                    "Phone": safe_get(user_data, "phone"),
                                    "Address": safe_get(user_data, "address"),
                                    "Tahasil": safe_get(user_data, "tahasil"),
                                    "District": safe_get(user_data, "district"),
                                    "Visit ID": safe_get(visit, "visit_id"),
                                    "Date & Time": visit_dt.strftime("%d/%m/%Y %I:%M:%S %p"),
                                    "Purpose": safe_get(visit, "purpose"),
                                    "Document": "Yes" if visit.get("document_pdf") else "No",
                                    "Status": safe_get(visit, "status", "unknown").capitalize(),
                                    "Entry Confirmed": "Yes" if visit.get("entry_confirmed", False) else "No",
                                    "Forwarded To": safe_get(visit, "forwarding_department", "-"),
                                    "Remark": safe_get(visit, "forwarding_note", "-"),
                                }
                                report_data.append(report_entry)
                        except ValueError:
                            print(f"Skipping visit due to invalid datetime format: UID {uid}, Visit {visit.get('visit_id', 'N/A')}")
                        except Exception as e:
                            print(f"Error processing visit row for report: UID {uid}, Visit {visit.get('visit_id', 'N/A')} - {e}")
        report_data.sort(key=lambda x: datetime.strptime(x["Date & Time"], "%d/%m/%Y %I:%M:%S %p"))
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Visitor Report"
        headers = [
            "Name",
            "Phone",
            "Address",
            "Tahasil",
            "District",
            "Visit ID",
            "Date & Time",
            "Purpose",
            "Document",
            "Status",
            "Entry Confirmed",
            "Forwarded To",
            "Remark",
        ]
        ws.append(headers)
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = openpyxl.styles.PatternFill(start_color="2C3E90", end_color="2C3E90", fill_type="solid")
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for entry in report_data:
            row_data = [entry.get(header, "N/A") for header in headers]
            ws.append(row_data)
        for col_idx, column_letter in enumerate(openpyxl.utils.get_column_letter(i) for i in range(1, len(headers) + 1)):
            max_length = 0
            column = ws[column_letter]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column_letter].width = max(15, min(adjusted_width, 50))
        excel_stream = io.BytesIO()
        wb.save(excel_stream)
        excel_stream.seek(0)
        filename = f"visitor_report_{start_date_str}.xlsx" if start_date == end_date else f"visitor_report_{start_date_str}_to_{end_date_str}.xlsx"
        return send_file(
            excel_stream,
            as_attachment=True,
            download_name=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    except Exception as e:
        print(f"CRITICAL: Error generating Excel report: {e}")
        return jsonify({"error": f"An unexpected server error occurred while generating the report: {str(e)}"}), 500


@app.route("/api/admin/delete_user", methods=["POST"])
@admin_login_required
def delete_user():
    data = request.get_json()
    if not data or "uid" not in data:
        return jsonify({"success": False, "message": "Missing UID in request body"}), 400
    uid_to_delete = data["uid"]
    try:
        record = get_face_item(uid_to_delete)
        if not record:
            return jsonify({"success": False, "message": f"Visitor UID '{uid_to_delete}' not found"}), 404
        if isinstance(record, dict) and "face_encodings" in record:
            for encoding_path in record.get("face_encodings", []):
                if encoding_path and isinstance(encoding_path, str) and s3_exists(encoding_path):
                    s3_delete(encoding_path)
        profile_prefix = s3_key(PROFILES_PREFIX, uid_to_delete) + "/"
        resp = s3.list_objects_v2(Bucket=BUCKET, Prefix=profile_prefix)
        for obj in resp.get("Contents", []):
            s3_delete(obj["Key"])
        face_data_prefix = s3_key(FACEDATA_PREFIX, uid_to_delete) + "/"
        resp = s3.list_objects_v2(Bucket=BUCKET, Prefix=face_data_prefix)
        for obj in resp.get("Contents", []):
            s3_delete(obj["Key"])
        if isinstance(record, dict) and "visitor" in record:
            for visit in record.get("visitor", []):
                if isinstance(visit, dict) and "document_pdf" in visit:
                    pdf_path = visit["document_pdf"]
                    if pdf_path and isinstance(pdf_path, str) and s3_exists(pdf_path):
                        s3_delete(pdf_path)
        delete_face_item(uid_to_delete)
        return jsonify({"success": True, "message": f"User '{uid_to_delete}' and associated files deleted successfully"})
    except Exception as e:
        print(f"Error deleting user {uid_to_delete}: {e}")
        traceback.print_exc()
        return jsonify({"success": False, "message": f"An server error occurred during user deletion: {str(e)}"}), 500


@app.route("/data/facedata/<uid>/<filename>")
def get_registration_image(uid, filename):
    key = s3_key(FACEDATA_PREFIX, uid, filename)
    if not s3_exists(key):
        alt_key = s3_key(FACEDATA_PREFIX, filename)
        if not s3_exists(alt_key):
            return jsonify({"success": False, "message": "Image not found"}), 404
        key = alt_key
    try:
        b = s3_get_bytes(key)
        return send_file(io.BytesIO(b), download_name=os.path.basename(key), mimetype="image/jpeg")
    except Exception:
        return jsonify({"success": False, "message": "Image not found"}), 404


@app.route("/api/visitor-status/<visit_id>", methods=["GET"])
def get_visitor_status(visit_id):
    if not visit_id:
        return jsonify({"success": False, "message": "Visit ID is required"}), 400
    try:
        data = scan_face_data()
        found_visitor_data = None
        target_visit = None
        visitor_uid = None
        for uid, user_data in data.items():
            if isinstance(user_data, dict) and "visitor" in user_data and isinstance(user_data["visitor"], list):
                for visit in user_data["visitor"]:
                    if isinstance(visit, dict) and visit.get("visit_id") == visit_id:
                        target_visit = visit
                        found_visitor_data = user_data
                        visitor_uid = uid
                        break
            if target_visit:
                break
        if not found_visitor_data or not target_visit:
            return jsonify({"success": False, "message": "Visitor ID not found"}), 404
        profile_image_url = None
        profile_imgs = found_visitor_data.get("profile_img", [])
        if profile_imgs and isinstance(profile_imgs, list) and profile_imgs[0]:
            profile_image_url = url_for("get_registration_image", uid=visitor_uid, filename=os.path.basename(profile_imgs[0]), _external=False)
        else:
            reg_images = found_visitor_data.get("images", [])
            if reg_images and isinstance(reg_images, list) and reg_images[0]:
                profile_image_url = url_for("get_registration_image", uid=visitor_uid, filename=os.path.basename(reg_images[0]), _external=False)
        response_data = {
            "success": True,
            "uid": visitor_uid,
            "name": found_visitor_data.get("name", "N/A"),
            "phone": found_visitor_data.get("phone", "N/A"),
            "tahasil": found_visitor_data.get("tahasil", "N/A"),
            "district": found_visitor_data.get("district", "N/A"),
            "visit_id": target_visit.get("visit_id", "N/A"),
            "datetime": target_visit.get("datetime", "N/A"),
            "purpose": target_visit.get("purpose", "N/A"),
            "status": target_visit.get("status", "N/A"),
            "forwarding_department": target_visit.get("forwarding_department", ""),
            "profile_image_url": profile_image_url,
            "face_encoding_paths": found_visitor_data.get("face_encodings", []),
            "image_paths": found_visitor_data.get("images", []),
            "dvn": target_visit.get("dvn"),
        }
        return jsonify(response_data)
    except Exception as e:
        print(f"Error in /api/visitor-status/{visit_id}: {e}")
        traceback.print_exc()
        return jsonify({"success": False, "message": f"An internal server error occurred: {str(e)}"}), 500


@app.route("/api/validate-visitor-face", methods=["POST"])
def validate_visitor_face():
    data = request.get_json()
    if not data or "uid" not in data or "image" not in data:
        return jsonify({"success": False, "message": "Missing uid or image data"}), 400
    uid = data["uid"]
    image_data_uri = data["image"]
    try:
        if "," not in image_data_uri:
            return jsonify({"success": False, "message": "Invalid image data format"}), 400
        header, encoded = image_data_uri.split(",", 1)
        image_bytes = base64.b64decode(encoded)
        nparr = np.frombuffer(image_bytes, np.uint8)
        captured_frame = cv2.imdecode(nparr, cv2.IMREAD_COLOR)
        if captured_frame is None:
            return jsonify({"success": False, "message": "Could not decode image"}), 400
        rgb_captured_frame = cv2.cvtColor(captured_frame, cv2.COLOR_BGR2RGB)
        user_data = get_face_item(uid)
        if not user_data:
            return jsonify({"success": False, "message": "User ID not found"}), 404
        known_encodings = []
        for path in user_data.get("face_encodings", []):
            try:
                b = s3_get_bytes(path)
                arr = np.load(io.BytesIO(b))
                known_encodings.append(arr)
            except Exception as e:
                print(f"Warning: Could not load encoding {path}: {e}")
        if not known_encodings:
            image_paths = user_data.get("images", [])
            if not image_paths:
                return jsonify({"success": False, "message": "No face data (encodings or images) found for this user to compare."}), 404
            for img_path_rel in image_paths:
                try:
                    b = s3_get_bytes(img_path_rel)
                    known_image = face_recognition.load_image_file(io.BytesIO(b))
                    encs = face_recognition.face_encodings(known_image, model=face_recognition_model)
                    if encs:
                        known_encodings.append(encs[0])
                except Exception as e:
                    print(f"Warning: Could not process image {img_path_rel} for encoding: {e}")
        if not known_encodings:
            return jsonify({"success": False, "message": "Could not load or generate any known face encodings for comparison."}), 404
        face_locations = face_recognition.face_locations(rgb_captured_frame, model=face_recognition_model)
        if not face_locations:
            return jsonify({"success": False, "message": "No face detected in the provided image."})
        captured_encodings = face_recognition.face_encodings(rgb_captured_frame, face_locations, model=face_recognition_model)
        if not captured_encodings:
            return jsonify({"success": False, "message": "Could not generate encoding from the detected face."})
        match_found = False
        for captured_encoding in captured_encodings:
            matches = face_recognition.compare_faces(known_encodings, captured_encoding, tolerance=0.45)
            if True in matches:
                match_found = True
                break
        if match_found:
            return jsonify({"success": True, "message": "Face validation successful."})
        else:
            return jsonify({"success": False, "message": "Face validation failed. No match found."})
    except Exception as e:
        print(f"Error during face validation for UID {uid}: {e}")
        traceback.print_exc()
        return jsonify({"success": False, "message": f"An internal server error occurred during face validation: {str(e)}"}), 500
    finally:
        cleanup_memory()


@app.route("/data/profiles/<uid>/<filename>")
def get_profile_image(uid, filename):
    key = s3_key(PROFILES_PREFIX, uid, filename)
    try:
        b = s3_get_bytes(key)
        # Serve as image/jpeg — if you have other types adapt as required
        return send_file(io.BytesIO(b), download_name=filename, mimetype="image/jpeg")
    except FileNotFoundError:
        return jsonify({"success": False, "message": "Profile image not found"}), 404
    except Exception as e:
        print(f"Error serving profile image {key}: {e}")
        return jsonify({"success": False, "message": "Error retrieving profile image"}), 500


if __name__ == "__main__":
    # threaded=True lets multiple requests be served in parallel
    app.run(host="0.0.0.0", port=5000, threaded=True)
